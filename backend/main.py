from fastapi import FastAPI, UploadFile, File, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import uvicorn
import shutil
import os
import uuid
import pypdfium2 as pdfium
from dotenv import load_dotenv
from ocr_engine import OCREngine
from llm_processor import LLMProcessor
from pydantic import BaseModel
from typing import List, Optional, Dict
import io
import csv
import anyio
from fastapi.responses import StreamingResponse
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load environment variables
load_dotenv()

app = FastAPI()

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize OCR Engine once (heavy model)
try:
    ocr_engine = OCREngine()
except Exception as e:
    print(f"Warning: OCR Engine failed to initialize: {e}")
    ocr_engine = None

# Ensure temp directory exists
os.makedirs("temp", exist_ok=True)
app.mount("/temp", StaticFiles(directory="temp"), name="temp")

@app.get("/status")
def get_status():
    """Returns the current status of the LLM processor"""
    status = getattr(LLMProcessor, '_current_status', "Ready")
    return {"status": status}

@app.get("/")
def read_root():
    return {
        "message": "2D Extraction API is running",
        "version": "EXPORT_FIX_V2",
        "endpoints": ["/upload", "/export/pdf", "/export/csv", "/test-pdf"]
    }

@app.get("/test-pdf")
async def test_pdf():
    """Diagnostic route to test PDF generation directly in browser"""
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("helvetica", "B", 16)
        pdf.cell(0, 10, "DIAGNOSTIC TEST PDF", ln=True, align="C")
        pdf_bytes = pdf.output()
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=test.pdf"}
        )
    except Exception as e:
        return {"error": str(e)}

# Helper for file logging
def log_debug(msg):
    try:
        with open("backend_debug.log", "a", encoding="utf-8") as f:
            import datetime
            timestamp = datetime.datetime.now().isoformat()
            f.write(f"[{timestamp}] {msg}\n")
    except:
        pass

@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...), 
    reference_file: Optional[UploadFile] = File(None),
    x_api_key: str = Header(None)
):
    log_debug(f"Received upload request. Primary: {file.filename}, Reference: {reference_file.filename if reference_file else 'None'}")

    if not file.filename.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg', '.dxf')):
        log_debug("Invalid file type rejected")
        raise HTTPException(status_code=400, detail="Invalid file type")

    # Initialize LLM Processor
    llm_processor = LLMProcessor(api_key=x_api_key)

    # Generate unique IDs
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1].lower()
    filename_base = f"{file_id}"
    file_path = os.path.join("temp", f"{filename_base}{ext}")

    try:
        # Save uploaded file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        log_debug(f"Saved primary file to {file_path}")

        # Save reference file if provided
        reference_image_path = None
        reference_filename = None
        if reference_file:
            ref_id = f"{file_id}_ref"
            ref_ext = os.path.splitext(reference_file.filename)[1].lower()
            ref_path = os.path.join("temp", f"{ref_id}{ref_ext}")
            with open(ref_path, "wb") as buffer:
                shutil.copyfileobj(reference_file.file, buffer)
            
            # Convert reference to image if PDF/DXF
            if ref_ext == '.pdf':
                pdf = pdfium.PdfDocument(ref_path)
                page = pdf[0]
                bitmap = page.render(scale=4)
                pil_image = bitmap.to_pil()
                reference_filename = f"{ref_id}.png"
                reference_image_path = os.path.join("temp", reference_filename)
                pil_image.save(reference_image_path)
            elif ref_ext == '.dxf':
                 from dxf_converter import convert_dxf_to_image
                 reference_filename = f"{ref_id}.png"
                 reference_image_path = os.path.join("temp", reference_filename)
                 convert_dxf_to_image(ref_path, reference_image_path)
            else:
                 reference_filename = f"{ref_id}{ref_ext}"
                 reference_image_path = ref_path
            
            log_debug(f"Saved reference image to {reference_image_path}")

        # Handle PDF to Image Conversion for Visuals
        display_image_filename = f"{filename_base}{ext}"
        
        if ext == '.pdf':
            log_debug("Converting PDF to image...")
            # Convert first page to image
            pdf = pdfium.PdfDocument(file_path)
            page = pdf[0]
            # Scale 4 = ~288 DPI (72 * 4), ensuring small text in title blocks is crisp for the LLM
            bitmap = page.render(scale=4) 
            pil_image = bitmap.to_pil()
            
            display_image_filename = f"{filename_base}.png"
            display_image_path = os.path.join("temp", display_image_filename)
            pil_image.save(display_image_path)
            log_debug(f"Converted PDF to image: {display_image_path}")
            
            # CRITICAL FIX: Analyze the IMAGE, not the PDF.
            # Use anyio to run the synchronous analyze_image in a separate thread
            analysis_result = await anyio.to_thread.run_sync(llm_processor.analyze_image, display_image_path)
            
        elif ext == '.dxf':
            log_debug("Converting DXF to image...")
            display_image_filename = f"{filename_base}.png"
            display_image_path = os.path.join("temp", display_image_filename)
            
            from dxf_converter import convert_dxf_to_image
            success = convert_dxf_to_image(file_path, display_image_path)
            
            if not success:
                raise Exception("Failed to convert DXF to image")
                
            log_debug(f"Converted DXF to image: {display_image_path}")
            # Use anyio to run the synchronous analyze_image in a separate thread
            analysis_result = await anyio.to_thread.run_sync(llm_processor.analyze_image, display_image_path)
            
        else:
            display_image_path = file_path
            # It's already an image
            # Use anyio to run the synchronous analyze_image in a separate thread
            analysis_result = await anyio.to_thread.run_sync(llm_processor.analyze_image, file_path)
        
        log_debug("Analysis complete. Returning response.")
        
        # URL for the frontend to confirm what to show
        image_url = f"http://localhost:8001/temp/{display_image_filename}"
        
        return {
            "filename": file.filename,
            "url": image_url,
            "features": analysis_result.get('features', []),
            "view_labels": analysis_result.get('view_labels', []),
            "metadata": analysis_result.get('metadata', {}),
            "primary_filename": display_image_filename,
            "reference_filename": reference_filename,
            "error": analysis_result.get('error')
        }

    except Exception as e:
        import traceback
        error_msg = f"Error processing file: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        log_debug(f"CRITICAL ERROR: {error_msg}")
        
        return {
            "filename": file.filename, 
            "error": str(e),
            "features": [],
            "metadata": {}
        }

@app.post("/compare")
async def compare_drawings(
    data: Dict,
    x_api_key: str = Header(None)
):
    primary_filename = data.get("primary_filename")
    reference_filename = data.get("reference_filename")
    
    log_debug(f"Received compare request for: {primary_filename} vs {reference_filename}")
    
    if not primary_filename or not reference_filename:
        raise HTTPException(status_code=400, detail="Missing filenames for comparison")
        
    primary_path = os.path.join("temp", primary_filename)
    reference_path = os.path.join("temp", reference_filename)
    
    if not os.path.exists(primary_path) or not os.path.exists(reference_path):
        raise HTTPException(status_code=404, detail="Files not found for comparison")
    
    llm_processor = LLMProcessor(api_key=x_api_key)
    
    try:
        # Use anyio to run comparison in a separate thread
        comparison_result = await anyio.to_thread.run_sync(
            llm_processor.compare_images, 
            primary_path, 
            reference_path
        )
        return comparison_result
    except Exception as e:
        log_debug(f"Error in compare_drawings: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

class ExportFeature(BaseModel):
    id: str
    type: str
    value: str
    view: str

class ExportRequest(BaseModel):
    features: List[ExportFeature]
    filename: str

class ExportMetadataRequest(BaseModel):
    metadata: Dict[str, str]
    filename: str

@app.post("/export/metadata/csv")
async def export_metadata_csv(data: ExportMetadataRequest):
    try:
        print(f"DEBUG: Starting Metadata CSV export for {data.filename}")
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Field", "Value"])
        
        print(f"DEBUG: Processing {len(data.metadata)} metadata items for {data.filename}")
        for key, value in data.metadata.items():
            writer.writerow([key, str(value).strip()])
        
        csv_content = output.getvalue()
        csv_bytes = csv_content.encode("utf-8-sig")
        
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{data.filename}_metadata.csv"'}
        )
    except Exception as e:
        import traceback
        print(f"METADATA CSV EXPORT ERROR: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/metadata/excel")
async def export_metadata_excel(data: ExportMetadataRequest):
    try:
        print(f"DEBUG: Starting Metadata Excel export for {data.filename}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Drawing Metadata"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Write Title
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Drawing Metadata: {data.filename}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Field", "Value"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        current_row = 4
        for key, value in data.metadata.items():
            ws.cell(row=current_row, column=1, value=key).border = border
            ws.cell(row=current_row, column=2, value=str(value)).border = border
            current_row += 1
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{data.filename}_metadata.xlsx"'}
        )
    except Exception as e:
        import traceback
        print(f"METADATA EXCEL EXPORT ERROR: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/metadata/pdf")
async def export_metadata_pdf(data: ExportMetadataRequest):
    try:
        print(f"DEBUG: Starting Metadata PDF export for {data.filename}")
        pdf = FPDF()
        pdf.add_page()
        
        # Use a standard font for metadata (Unicode support if possible)
        has_bold = False
        try:
            # Look for common Windows fonts
            font_paths = [
                ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
                ("C:/Windows/Fonts/consola.ttf", "C:/Windows/Fonts/consolab.ttf"),
                ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
            ]
            
            selected_fonts = next(((reg, bld) for reg, bld in font_paths if os.path.exists(reg)), (None, None))
            reg_path, bld_path = selected_fonts

            if reg_path:
                pdf.add_font("CustomUnicode", "", reg_path)
                if bld_path and os.path.exists(bld_path):
                    pdf.add_font("CustomUnicode", "B", bld_path)
                    has_bold = True
                
                pdf.set_font("CustomUnicode", size=12)
                print(f"DEBUG: Using Custom Unicode font for metadata (Bold: {has_bold})")
            else:
                pdf.set_font("Helvetica", size=12)
                has_bold = True # Built-in Helvetica always has bold
        except Exception as e:
            print(f"DEBUG: Font registration failed: {e}")
            pdf.set_font("Helvetica", size=12)
            has_bold = True

        # Helper to set bold style safely
        def set_safe_bold(size):
            if has_bold:
                pdf.set_font(size=size, style='B')
            else:
                pdf.set_font(size=size, style='')

        # Title
        set_safe_bold(16)
        pdf.cell(200, 10, txt=f"Drawing Metadata Report", ln=True, align='C')
        set_safe_bold(12)
        pdf.cell(200, 10, txt=f"Designation: {data.filename}", ln=True, align='C')
        pdf.ln(10)
        
        # Table Headers
        pdf.set_fill_color(220, 220, 220)
        set_safe_bold(10)
        pdf.cell(60, 10, "Field", border=1, fill=True)
        pdf.cell(130, 10, "Value", border=1, fill=True, ln=True)
        
        # Table Content
        pdf.set_font(size=10, style='')
        for key, value in data.metadata.items():
            # Handle potential multi-line values
            key_text = str(key)
            val_text = str(value)
            
            # Use multi_cell for automatic wrapping
            # We need to manually sync heights of adjacent cells
            # Find how many lines each would take
            # fpdf2 doesn't have a simple "get_multi_cell_height", but we can use multi_cell with split_only=True
            # However, for simplicity, we'll just use the approach of finding max height
            
            x = pdf.get_x()
            y = pdf.get_y()
            
            # Print Key
            pdf.multi_cell(60, 10, key_text, border=1)
            y_key = pdf.get_y()
            
            # Reset to print Val next to it
            pdf.set_y(y)
            pdf.set_x(x + 60)
            pdf.multi_cell(130, 10, val_text, border=1)
            y_val = pdf.get_y()
            
            # Move to next row
            pdf.set_y(max(y_key, y_val))
            pdf.set_x(x)

        pdf_bytes = pdf.output()
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{data.filename}_metadata.pdf"'}
        )
    except Exception as e:
        import traceback
        print(f"METADATA PDF EXPORT ERROR: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/csv")
async def export_csv(data: ExportRequest):
    try:
        print(f"DEBUG: Starting CSV export for {data.filename} with {len(data.features)} features")
        log_debug(f"Starting CSV export for: {data.filename}")
        
        # Mapping to normalize diameter variants to a single standard Ø
        # We restore true Unicode symbols for GD&T features
        SYMBOL_MAP = {
            '\u2300': '\u00d8', # ⌀ -> Ø
            '\u2304': '\u00d8', # ⌀ -> Ø
            '\u03d5': '\u00d8', # ϕ (Greek Phi) -> Ø
            '\u03a6': '\u00d8', # Φ (Greek Phi) -> Ø
            '\u03c6': '\u00d8', # φ (Greek Phi) -> Ø
            '\u00f8': '\u00d8', # lowercase ø -> Ø
        }

        def clean_val(v):
            if not v: return ""
            res = str(v).strip()
            for char, repl in SYMBOL_MAP.items():
                res = res.replace(char, repl)
            return res

        # Comma (,) is the standard CSV delimiter.
        # UTF-8 with BOM (utf-8-sig) ensures Excel/CATIA auto-detects symbols correctly.
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["View", "ID", "Type", "Value"], delimiter=',')
        writer.writeheader()
        
        def sort_key(f):
            try: numeric_id = int(f.id)
            except: numeric_id = 999
            return (f.view or "General", f.type or "Other", numeric_id)

        sorted_features = sorted(data.features, key=sort_key)
        
        for f in sorted_features:
            writer.writerow({
                "View": clean_val(f.view or "General"),
                "ID": clean_val(f.id),
                "Type": clean_val(f.type or "Dimension"),
                "Value": clean_val(f.value)
            })
        
        # Use UTF-8 with BOM (utf-8-sig) for perfect Unicode detection in Windows apps
        csv_content = output.getvalue()
        csv_bytes = csv_content.encode("utf-8-sig") 
        
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{data.filename}.csv"',
                "Content-Type": "text/csv; charset=utf-8"
            }
        )
    except Exception as e:
        import traceback
        log_debug(f"CSV EXPORT ERROR: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/export/excel")
async def export_excel(data: ExportRequest):
    try:
        print(f"DEBUG: Starting Excel export for {data.filename}")
        log_debug(f"Starting Excel export for: {data.filename}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extraction Results"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        view_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Write Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Extraction Results: {data.filename}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["View", "ID", "Type", "Value"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        # Group by View
        views = {}
        for f in data.features:
            view = f.view or "General"
            if view not in views:
                views[view] = []
            views[view].append(f)
            
        current_row = 4
        for view_name, features in sorted(views.items()):
            # View Row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1, value=f"VIEW: {view_name.upper()}")
            cell.font = Font(bold=True)
            cell.fill = view_fill
            cell.border = border
            current_row += 1
            
            def try_int(v):
                try: return int(v)
                except: return 999
            
            sorted_f = sorted(features, key=lambda x: (x.type, try_int(x.id)))
            
            for f in sorted_f:
                ws.cell(row=current_row, column=1, value=f.view).border = border
                ws.cell(row=current_row, column=2, value=f.id).border = border
                ws.cell(row=current_row, column=3, value=f.type).border = border
                ws.cell(row=current_row, column=4, value=f.value).border = border
                current_row += 1
            current_row += 1
            
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{data.filename}.xlsx"'}
        )
    except Exception as e:
        import traceback
        err_detail = f"EXCEL EXPORT ERROR: {str(e)}\n{traceback.format_exc()}"
        print(err_detail)
        log_debug(err_detail)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/pdf")
async def export_pdf(data: ExportRequest):
    try:
        print(f"DEBUG: Starting PDF export for {data.filename}")
        log_debug(f"Starting PDF export for: {data.filename}")
        
        # Use fpdf2 which is required for unicode
        pdf = FPDF()
        pdf.add_page()
        
        # Try to add a Unicode font if on Windows
        font_path = "C:/Windows/Fonts/arial.ttf"
        has_unicode = False
        if os.path.exists(font_path):
            try:
                # Add font and set it as the current font
                pdf.add_font("Arial", "", font_path)
                pdf.set_font("Arial", "", 12)
                has_unicode = True
                print("DEBUG: Using Arial Unicode font")
            except Exception as fe:
                print(f"DEBUG: Font load failed: {fe}")
                pdf.set_font("helvetica", "", 12)
        else:
            print("DEBUG: Arial.ttf not found, using helvetica fallback")
            pdf.set_font("helvetica", "", 12)
        
        # Title
        pdf.set_font(pdf.font_family, "B" if not has_unicode else "", 16)
        pdf.cell(0, 10, f"Extraction Results: {data.filename}", ln=True, align="C")
        pdf.ln(5)
        
        # Group by View
        views = {}
        for f in data.features:
            view = f.view or "General"
            if view not in views:
                views[view] = {}
            
            typ = f.type or "Other"
            if typ not in views[view]:
                views[view][typ] = []
            views[view][typ].append(f)
            
        for view_name, type_groups in sorted(views.items()):
            # View Header
            pdf.set_font(pdf.font_family, "B" if not has_unicode else "", 12)
            pdf.set_fill_color(230, 240, 255)
            pdf.cell(0, 10, f" VIEW: {view_name.upper()}", ln=True, fill=True)
            pdf.ln(2)
            
            for typ_name, features in sorted(type_groups.items()):
                # Type Sub-header
                pdf.set_font(pdf.font_family, "B" if not has_unicode else "", 10)
                pdf.set_text_color(100, 100, 100)
                pdf.cell(0, 8, f"  {typ_name}", ln=True)
                pdf.set_text_color(0, 0, 0)
                
                # Table Header
                pdf.set_font(pdf.font_family, "B" if not has_unicode else "", 9)
                pdf.cell(20, 7, "ID", border=1, align="C")
                pdf.cell(0, 7, "Value", border=1, ln=True)
                
                # Table Rows
                pdf.set_font(pdf.font_family, "", 9)
                for f in features:
                    val = f.value
                    if not has_unicode:
                        # Replace symbols for standard Helvetica (latin-1)
                        val = val.replace('\u00d8', 'Diam. ').replace('\u00b1', '+/-').replace('\u00b0', ' deg')
                        val = val.encode('latin-1', 'replace').decode('latin-1')
                    
                    pdf.cell(20, 7, f.id, border=1, align="C")
                    pdf.cell(0, 7, f" {val}", border=1, ln=True)
                
                pdf.ln(3)
            pdf.ln(2)

        # In fpdf2, output() returns bytes by default
        pdf_bytes = pdf.output()
        print("DEBUG: PDF generated successfully")
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{data.filename}.pdf"',
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        import traceback
        err_detail = f"PDF EXPORT ERROR: {str(e)}\n{traceback.format_exc()}"
        print(err_detail)
        log_debug(err_detail)
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)
